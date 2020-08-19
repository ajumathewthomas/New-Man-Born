#Automation of Wifi Audit using Python
# Authors : AMT, GAK


from reportlab.pdfgen import canvas
import os
import openpyxl
from reportlab.lib.pagesizes import letter,A4,A3
from reportlab.lib.pagesizes import landscape,portrait
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
import gmplot

os.chdir(r'H:\war driving\report')
wb = openpyxl.load_workbook('wifi.xlsx')
ws = wb.active
filename = 'Wifi'+'_'+'Audit'+'_'+'Report'+'.pdf'


def generate_report():
    title = 'Wi-Fi Audit Report'
    wifi_name = ws['B3'].value
    wifi_auth = ws['C3'].value
    wifi_enc = ws['D3'].value
    wifi_channel = ws['F3'].value
    wifi_lat = ws['H3'].value
    wifi_lon = ws['I3'].value
    wifi_mac = ws['A3'].value
    wifi_rssi = ws['G3'].value
    wifi_am = ws['K3'].value
    pdf = canvas.Canvas(filename, pagesize=portrait(A4))
    pdf.setTitle(title)
    pdf.setFont('Helvetica-Bold', size=16, leading=None)
    pdf.drawString(270, 770, title)
    pdf.line(270,760,410,760)
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(100, 700, "SSID Name:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(170, 700, wifi_name)
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(250, 700, "Authentication:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(340, 700, wifi_auth)
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(420, 700, "Encryption:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(490, 700, wifi_enc)
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(100, 650, "Channel:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(155, 650, str(wifi_channel))
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(250, 650, "Latitude:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(300, 650, str(wifi_lat))
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(420, 650, "Longitude:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(485, 650, str(wifi_lon))
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(100, 600, "BSSID:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(147, 600, str(wifi_mac))
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(275, 600, "RSSI:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(310, 600, str(wifi_rssi))
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(420, 600, "Accuracy:")
    pdf.setFont('Helvetica', size=12, leading=None)
    pdf.drawString(485, 600, str(wifi_am))
    pdf.setFont('Helvetica-Bold', size=12, leading=None)
    pdf.drawString(100, 550, "Remarks:")
    wb1 = openpyxl.load_workbook("policy.xlsx")
    ws1 = wb1.active
    w1 = ws1['B4'].value
    w2 = ws1['B5'].value
    w3 = ws1['B6'].value
    w4 = ws1['B7'].value
    w5 = ws1['B8'].value
    w6 = ws1['B9'].value
    w7 = ws1['B10'].value
    pdf.setFont('Helvetica', size=8, leading=None)

    pdf.drawString(90, 525, "1." + str(w1))
    pdf.setFont('Helvetica', size=8, leading=None)
    pdf.drawString(90, 510, "2." + str(w2))
    pdf.setFont('Helvetica', size=8, leading=None)
    pdf.drawString(90, 495, "3." + str(w3))
    pdf.setFont('Helvetica', size=8, leading=None)
    pdf.drawString(90, 480, "4." + str(w4))
    pdf.setFont('Helvetica', size=8, leading=None)
    pdf.drawString(90, 465, "5." + str(w5))
    pdf.setFont('Helvetica', size=8, leading=None)
    pdf.drawString(90, 450, "6." + str(w6))
    pdf.setFont('Helvetica', size=8, leading=None)
    pdf.drawString(90, 435, "7." + str(w7))
    wb1.close()
    pdf.showPage()
    pdf.save()

def googPlot():
    #nonlocal  gmapOne
    latt =10.5276
    long =76.2144
    gmapOne = gmplot.GoogleMapPlotter(ws['H3'].value,ws['I3'].value, zoom=15)
    gmapOne.coloricon = "http://www.googlemapsmarkers.com/v1/%s/"
    lat = ws['H3'].value
    lon = ws['I3'].value
    gmapOne.marker(lat,lon,'red')
    #colors = ['#FFFF00','#8B0000','#FFFF00']
    #gmapOne.scatter(lat,lon,size=10,color='8B0000',marker=False)

    gmapOne.draw('H:\\war driving\\report\\map.html')




for i in range(0,3):
     if ws.cell(row=i + 1, column=4).value == 'None':

         generate_report()
         googPlot()


         fromEmail ="Your Email ID"
         fromPass ="Your password"
         toEmail ="To Email"
         toCC = " CC email"
         subject = 'Wi-Fi Audit Report' + ' ' + str(datetime.date(datetime.now()))
         msg = MIMEMultipart()
         msg['Subject'] = subject
         msg['From'] = fromEmail
         msg['To'] = toEmail
         msg['Cc'] = toCC
         toAddr = [toEmail] + [toCC]
         body = ''' Body of the message'''
         msg.attach(MIMEText(body, 'plain'))
         filepath = "path"
         filepath1 = "path"
         filename = os.path.basename(filepath)
         attachment = open(filepath, "rb")
         part = MIMEBase('application', 'octet-stream')
         part.set_payload(attachment.read())
         encoders.encode_base64(part)
         part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
         filenamet = os.path.basename(filepath1)
         attachmentt = open(filepath1, "rb")
         partt = MIMEBase('application', 'octet-stream')
         partt.set_payload(attachmentt.read())
         encoders.encode_base64(partt)
         partt.add_header('Content-Disposition', "attachment; filename= %s" % filenamet)
         msg.attach(part)
         msg.attach(partt)
         session = smtplib.SMTP('smtp.office365.com', 587)  # use gmail with port
         session.starttls()  # enable security

         session.login(fromEmail, fromPass)  # login with mail_id and password

         session.sendmail(fromEmail, toAddr, msg.as_string())
         session.quit()


def drawMyRuler(pdf):
    pdf.drawString(100, 810, 'x100')
    pdf.drawString(200, 810, 'x200')
    pdf.drawString(300, 810, 'x300')
    pdf.drawString(400, 810, 'x400')
    pdf.drawString(500, 810, 'x500')

    pdf.drawString(10, 100, 'y100')
    pdf.drawString(10, 200, 'y200')
    pdf.drawString(10, 300, 'y300')
    pdf.drawString(10, 400, 'y400')
    pdf.drawString(10, 500, 'y500')
    pdf.drawString(10, 600, 'y600')
    pdf.drawString(10, 700, 'y700')
    pdf.drawString(10, 800, 'y800')


#drawMyRuler(pdf)







